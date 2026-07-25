"""Importar lista de precios desde Excel (GERENCIA only).

Permite a GERENCIA subir un Excel con precios, validarlos contra el catálogo
existente (el código debe existir), ver una vista previa y confirmar la
actualización masiva.

Reglas:
- Solo se actualizan productos cuyo Código ya exista en el sistema.
- Códigos desconocidos se marcan como error.
- No se modifica stock, sede ni material_tipo.
- No se eliminan productos.
- Los precios deben ser números no negativos (vacío = sin cambio).
- Los códigos deben ser únicos dentro del archivo cargado.
"""

from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation

import openpyxl
import pandas as pd
import streamlit as st

from ecomajes import config, db

# Columnas de la plantilla (en orden)
_TEMPLATE_COLS = [
    "Código",
    "Descripción",
    "Familia",
    "Unidad",
    "Precio en dólares",
    "P1",
    "P2",
    "P3",
    "Venta mínima",
    "Venta oficial",
    "Venta por 3 metros",
    "Venta por metro",
    "Observaciones",
]

# Columna de plantilla -> campo de la tabla prices
_PRICE_MAP = {
    "Precio en dólares": "precio",
    "P1": "p1",
    "P2": "p2",
    "P3": "p3",
    "Venta mínima": "precio_minimo",
    "Venta oficial": "venta_oficial",
    "Venta por 3 metros": "venta_3m",
    "Venta por metro": "venta_metro",
}

# Sinónimos de cabecera (key en minúsculas) -> nombre canónico de la plantilla
_ALIASES: dict[str, str] = {
    "código": "Código",
    "codigo": "Código",
    "cod": "Código",
    "cod.": "Código",
    "descripcion": "Descripción",
    "descripción": "Descripción",
    "desc": "Descripción",
    "familia": "Familia",
    "unidad": "Unidad",
    "und": "Unidad",
    "u.m.": "Unidad",
    "precio en dólares": "Precio en dólares",
    "precio en dolares": "Precio en dólares",
    "precio usd": "Precio en dólares",
    "precio": "Precio en dólares",
    "p1": "P1",
    "p2": "P2",
    "p3": "P3",
    "venta mínima": "Venta mínima",
    "venta minima": "Venta mínima",
    "precio mínimo": "Venta mínima",
    "precio minimo": "Venta mínima",
    "venta oficial": "Venta oficial",
    "venta por 3 metros": "Venta por 3 metros",
    "venta 3m": "Venta por 3 metros",
    "venta por metro": "Venta por metro",
    "venta metro": "Venta por metro",
    "observaciones": "Observaciones",
    "obs": "Observaciones",
    "notas": "Observaciones",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell(v) -> str:
    """Normaliza una celda a texto limpio ('' para NaN/None)."""
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    try:
        if pd.isna(v):
            return ""
    except (TypeError, ValueError):
        pass
    s = str(v).strip()
    # Floats enteros vienen como "10.0" → "10"
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def _parse_price(v) -> tuple[Decimal | None, str | None]:
    """Devuelve (decimal, mensaje_error). Vacío → (None, None). Negativo → error."""
    s = _cell(v)
    if not s:
        return None, None
    try:
        d = Decimal(s.replace(",", "."))
    except InvalidOperation:
        return None, f"'{s}' no es un número válido"
    if d < 0:
        return None, f"{d} es negativo"
    return d, None


def _build_template() -> bytes:
    """Genera la plantilla Excel con cabeceras y una fila de ejemplo."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Precios"

    # Cabeceras en negrita
    ws.append(_TEMPLATE_COLS)
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1565C0")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18

    # Fila de ejemplo (se puede borrar)
    ws.append([
        "COD-001",        # Código
        "Tubo cuadrado 40x40",  # Descripción
        "Acero",          # Familia
        "ML",             # Unidad
        12.50,            # Precio en dólares
        11.00,            # P1
        10.50,            # P2
        10.00,            # P3
        6.00,             # Venta mínima
        13.00,            # Venta oficial
        18.00,            # Venta por 3 metros
        6.00,             # Venta por metro
        "",               # Observaciones
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_upload(file) -> list[dict]:
    """Lee el Excel subido y devuelve filas crudas con las columnas de la plantilla."""
    df = pd.read_excel(file, dtype=str)

    # Normalizar cabeceras
    rename: dict = {}
    for col in df.columns:
        canon = _ALIASES.get(str(col).strip().lower())
        if canon:
            rename[col] = canon
    df = df.rename(columns=rename)

    rows: list[dict] = []
    for _, r in df.iterrows():
        rows.append({col: r.get(col, "") for col in _TEMPLATE_COLS})
    return rows


def _validate(
    raw_rows: list[dict],
    precio_index: dict[str, dict],
) -> tuple[list[dict], list[dict]]:
    """Valida filas crudas contra el índice de productos existentes.

    Devuelve (filas_válidas, filas_con_error).
    Cada fila válida tiene 'product_id' inyectado y los precios como Decimal|None.
    Cada fila de error tiene '_fila' (nro de fila en Excel) y '_error' (descripción).
    """
    valid: list[dict] = []
    errors: list[dict] = []
    seen_codes: set[str] = set()

    for i, row in enumerate(raw_rows, start=2):  # fila 2 = primera de datos
        codigo_raw = _cell(row.get("Código", ""))
        codigo = codigo_raw.upper()

        # Validación 1: código vacío
        if not codigo:
            errors.append({**row, "_fila": i, "_error": "Código vacío — obligatorio"})
            continue

        # Validación 2: código duplicado en el archivo
        if codigo in seen_codes:
            errors.append({
                **row,
                "_fila": i,
                "_error": f"Código duplicado en el archivo: {codigo}",
            })
            continue
        seen_codes.add(codigo)

        # Validación 3: código no existe en el sistema
        if codigo not in precio_index:
            errors.append({
                **row,
                "_fila": i,
                "_error": f"Código no existe en el sistema: {codigo}",
            })
            continue

        # Validación 4: precios numéricos y no negativos
        row_errors: list[str] = []
        parsed_prices: dict = {}
        for col, field in _PRICE_MAP.items():
            val, err = _parse_price(row.get(col, ""))
            if err:
                row_errors.append(f"{col}: {err}")
            else:
                parsed_prices[field] = val

        if row_errors:
            errors.append({**row, "_fila": i, "_error": "; ".join(row_errors)})
            continue

        # Fila válida
        product_id = precio_index[codigo]["product_id"]
        valid.append({
            "product_id": product_id,
            "codigo": codigo,
            "descripcion": _cell(row.get("Descripción", "")) or None,
            "familia": _cell(row.get("Familia", "")) or None,
            "unidad": _cell(row.get("Unidad", "")) or None,
            **parsed_prices,
        })

    return valid, errors


# ---------------------------------------------------------------------------
# View
# ---------------------------------------------------------------------------

def render(ctx: dict) -> None:
    st.header(ctx["title"])
    st.caption(ctx["breadcrumb"])

    # Guardia: solo GERENCIA
    if ctx.get("role") != config.ROLE_GERENCIA:
        st.error("⛔ Esta función es exclusiva de GERENCIA.")
        return

    st.write(
        "Descarga la plantilla, completa los precios y sube el archivo "
        "para actualizar la lista de precios de forma masiva."
    )

    # ── Paso 1: Plantilla ──────────────────────────────────────────────────
    st.subheader("1 · Descargar plantilla")
    st.download_button(
        label="⬇️ Descargar plantilla Excel",
        data=_build_template(),
        file_name="plantilla_precios_ecomajes.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Descarga la plantilla con las columnas requeridas.",
    )

    st.divider()

    # ── Paso 2: Subir archivo ──────────────────────────────────────────────
    st.subheader("2 · Subir archivo Excel")
    uploaded = st.file_uploader(
        "Selecciona el archivo Excel con los precios actualizados",
        type=["xlsx", "xls"],
        key="importar_precios_file",
    )
    if not uploaded:
        st.info("Sube el archivo Excel para continuar.")
        return

    # Cargar catálogo de productos con precio desde DB (todas las sedes)
    try:
        all_rows = db.list_prices(include_all_sedes=True)
    except Exception as exc:
        st.error(f"Error al cargar el catálogo: {exc}")
        return

    # Índice código (upper) → fila del catálogo
    precio_index: dict[str, dict] = {}
    for p in all_rows:
        cod = (p.get("codigo") or "").strip().upper()
        if cod:
            precio_index[cod] = p

    if not precio_index:
        st.warning(
            "No hay productos con código asignado en el sistema. "
            "Registra productos con código en el catálogo antes de importar precios."
        )
        return

    # Parsear Excel subido
    try:
        raw_rows = _parse_upload(uploaded)
    except Exception as exc:
        st.error(f"No se pudo leer el archivo Excel: {exc}")
        return

    if not raw_rows:
        st.warning("El archivo no contiene filas de datos.")
        return

    # ── Paso 3: Validación y vista previa ─────────────────────────────────
    valid_rows, error_rows = _validate(raw_rows, precio_index)

    st.divider()
    st.subheader("3 · Vista previa y validación")

    c1, c2, c3 = st.columns(3)
    c1.metric("📋 Filas leídas", len(raw_rows))
    c2.metric("✅ Válidos", len(valid_rows))
    c3.metric("❌ Errores", len(error_rows))

    # Tabla de errores
    if error_rows:
        with st.expander(
            f"⚠️ Ver {len(error_rows)} error(es) — estos registros NO se importarán",
            expanded=(len(valid_rows) == 0),
        ):
            err_df = pd.DataFrame([
                {
                    "Fila (Excel)": r["_fila"],
                    "Código": _cell(r.get("Código", "")),
                    "Descripción": _cell(r.get("Descripción", "")),
                    "Error": r["_error"],
                }
                for r in error_rows
            ])
            st.dataframe(err_df, use_container_width=True, hide_index=True)

    if not valid_rows:
        st.error("No hay registros válidos para importar. Corrige los errores y vuelve a subir.")
        return

    # Vista previa de cambios válidos
    _PREVIEW_COLS = {
        "codigo": "Código",
        "descripcion": "Descripción",
        "familia": "Familia",
        "unidad": "Unidad",
        "precio": "Precio USD",
        "p1": "P1",
        "p2": "P2",
        "p3": "P3",
        "precio_minimo": "Venta mínima",
        "venta_oficial": "Venta oficial",
        "venta_3m": "Venta 3m",
        "venta_metro": "Venta/metro",
    }
    preview_data = [
        {
            label: (
                str(row[field]) if row.get(field) is not None else "—"
            )
            for field, label in _PREVIEW_COLS.items()
        }
        for row in valid_rows
    ]
    st.write(f"**Productos que serán actualizados ({len(valid_rows)}):**")
    st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)

    # ── Paso 4: Confirmar ──────────────────────────────────────────────────
    st.divider()
    st.subheader("4 · Confirmar importación")

    if error_rows:
        st.warning(
            f"{len(error_rows)} registro(s) con errores serán ignorados. "
            f"Solo se importarán los {len(valid_rows)} válidos."
        )

    st.info(
        "✔ Se actualizarán: descripción, familia, unidad y precios.  \n"
        "✘ NO se modificará: stock, sede, tipo de material.  \n"
        "✘ NO se eliminarán productos."
    )

    if st.button(
        f"✅ Confirmar e importar {len(valid_rows)} producto(s)",
        type="primary",
    ):
        try:
            # Preparar filas para prices table (precio_sugerido espeja precio)
            price_rows = []
            for row in valid_rows:
                price_rows.append({
                    "product_id": row["product_id"],
                    "codigo": row["codigo"],
                    "descripcion": row.get("descripcion"),
                    "unidad": row.get("unidad"),
                    "precio": row.get("precio"),
                    "precio_sugerido": row.get("precio"),  # espejo obligatorio
                    "p1": row.get("p1"),
                    "p2": row.get("p2"),
                    "p3": row.get("p3"),
                    "precio_minimo": row.get("precio_minimo"),
                    "venta_oficial": row.get("venta_oficial"),
                    "venta_3m": row.get("venta_3m"),
                    "venta_metro": row.get("venta_metro"),
                })

            n = db.save_prices(price_rows)

            # Actualizar familia en tabla products
            familia_updates = [
                (row["product_id"], row["familia"])
                for row in valid_rows
                if row.get("familia")
            ]
            if familia_updates:
                db.update_product_familia_batch(familia_updates)

        except Exception as exc:
            st.error(f"Error al guardar los precios: {exc}")
            return

        db.log_audit(
            db.AUDIT_PRICE_CHANGED,
            "Importar Precios Excel",
            detalle=(
                f"{n} producto(s) actualizados vía importación Excel. "
                f"Errores omitidos: {len(error_rows)}."
            ),
            usuario_rol=ctx["usuario_rol"],
            sede=ctx.get("sede"),
        )
        st.success(f"✅ Importación completada: {n} producto(s) actualizados correctamente.")
        st.rerun()
